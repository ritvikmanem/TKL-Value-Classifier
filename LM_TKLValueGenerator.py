from openpyxl import load_workbook
import os
import sys

def TKLRetrieval(excel_file, output_dir):
    #function to extract and append unique values from specific columns in an Excel file to text files
    try:
        workbook = load_workbook(filename=excel_file)  #load the Excel workbook
        sheet = workbook.active  #select the active worksheet
    except Exception as e:
        print(f"Error loading workbook {excel_file}: {e}")  #log any loading errors
        return

    #define paths for output files to store unique values
    type_file_path = os.path.join(output_dir, 'projectTypeTKLValues.txt')
    doc_type_file_path = os.path.join(output_dir, 'documentTypeTKLValues.txt')
    metadata_file_path = os.path.join(output_dir, 'metadataReviewedTKLValues.txt')

    def read_existing_values(file_path):
        #function to read existing values from a file and return them as a set
        try:
            if os.path.exists(file_path):  #check if the file exists
                with open(file_path, 'r') as file:
                    return set(line.strip().lower().title() for line in file)  #normalize and store unique values
        except Exception as e:
            print(f"Error reading {file_path}: {e}")  #log file read errors
        return set()

    #read existing values to avoid duplicates
    type_set = read_existing_values(type_file_path)
    doc_type_set = read_existing_values(doc_type_file_path)
    metadata_set = read_existing_values(metadata_file_path)

    try:
        #open output files in append mode
        with open(type_file_path, 'a') as typeFile, open(doc_type_file_path, 'a') as docTypeFile, open(metadata_file_path, 'a') as metadataFile:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
                #iterate through rows starting from the second (to skip headers)

                if row[9]:  #process column 10 (index 9)
                    value = row[9].strip().lower().title()
                    if value not in type_set:  # Avoid duplicates
                        typeFile.write(f'{value}\n')  #write unique value to file
                        type_set.add(value)

                if row[11]:  #process column 12 (index 11)
                    value = row[11].strip().lower().title()
                    if value not in doc_type_set:  #avoid duplicates
                        docTypeFile.write(f'{value}\n')  #write unique value to file
                        doc_type_set.add(value)

                if row[17]:  #process column 18 (index 17)
                    value = row[17].strip().lower().title()
                    if value not in metadata_set:  #avoid duplicates
                        metadataFile.write(f'{value}\n')  #write unique value to file
                        metadata_set.add(value)
    except Exception as e:
        print(f"Error writing to files: {e}")  #log file write errors

def process_directory(input_dir, output_dir):
    #function to process all subdirectories in the input directory
    if not input_dir or not output_dir:
        return  #exit if directories are not valid
    try:
        for subdir in os.listdir(input_dir):  #list subdirectories in input directory
            subdir_path = os.path.join(input_dir, subdir)
            if os.path.isdir(subdir_path):  #check if path is a directory
                excel_files = [f for f in os.listdir(subdir_path) if f.endswith('Full Index.xlsx')]
                #filter for Excel files matching specific naming pattern
                if excel_files:
                    excel_file = os.path.join(subdir_path, excel_files[0])  #select the first matching file
                    os.makedirs(output_dir, exist_ok=True)  #ensure output directory exists
                    TKLRetrieval(excel_file, output_dir)  #process the Excel file
    except Exception as e:
        print(f"Error processing directory {input_dir}: {e}")  #log directory processing errors

if __name__ == '__main__':
    #main entry point of the program
    if len(sys.argv) != 3:
        print("Usage: python script.py <input_directory> <output_directory>")
        sys.exit(1)

    input_dir = sys.argv[1]  #input directory from command-line argument
    output_dir = sys.argv[2]  #output directory from command-line argument
    process_directory(input_dir, output_dir)  #process all subdirectories
    input("Program Completed")  #pause for user acknowledgment
